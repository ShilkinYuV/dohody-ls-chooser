import openpyxl
from PyQt5.QtGui import QFont
from openpyxl.styles import Alignment,Font

from openpyxl.workbook.workbook import Workbook
from MainForm import Ui_MainWindow
from AboutWindow import AboutWindows
from PyQt5.QtWidgets import QAbstractItemView, QDesktopWidget, QDialog, QFileDialog, QLabel, QTableWidgetItem, QHeaderView, QMessageBox, QWidget
from PyQt5 import QtWidgets, QtCore
import os

from classes.ReadExcel import ReadExcel

class MainWindow(QtWidgets.QMainWindow):
    def __init__(self) -> None:
        super(MainWindow,self).__init__()
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        self.aboutForm = None
        self.ui.pbtn_add_report.clicked.connect(self.open_file)
        self.ui.action_save_to_excel.triggered.connect(self.ExportResultsToExcel)
        self.ui.tableWidget.setRowCount(1)
        self.ui.tableWidget.setColumnCount(5)
        self.ui.tableWidget.horizontalHeader().setVisible(False)
        self.ui.tableWidget.verticalHeader().setVisible(False)
        self.ui.tableWidget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.ui.action_about_program.triggered.connect(self.OpenAbout)
        self.font = QFont()
        self.font.setBold(True)
        self.font.setPixelSize(14)
        self.font_t = QFont()
        self.font_t.setPixelSize(14)

        self.writeTBheader()

    def open_file(self):
        self.ui.statusbar.showMessage('')
        self.ui.tableWidget.clear()
        self.writeTBheader()
        self.filename, _ = QFileDialog.getOpenFileName(
            None, 'Открыть', os.path.join(os.path.join(
                os.environ['USERPROFILE']), 'Desktop'),
            'Excel Files(*.xls);;Excel Files(*.xlsx)')


        if self.filename in "('', '')":
            self.ui.statusbar.showMessage('Файл не выбран')

        else:
            self.ui.pbtn_add_report.setEnabled(False)
            self.my_thread = ReadExcel(my_window=self)
            self.my_thread.appendText.connect(self.appendText)
            self.my_thread.start()

    def writeTBheader(self):
        self.ui.tableWidget.setRowCount(1)
        newItem = QTableWidgetItem("Лицевой счет")
        newItem.setFont(self.font)
        newItem.setTextAlignment(QtCore.Qt.AlignCenter)
        self.ui.tableWidget.setItem(0, 0, newItem)

        newItem = QTableWidgetItem("Возвраты")
        newItem.setFont(self.font)
        newItem.setTextAlignment(QtCore.Qt.AlignCenter)
        self.ui.tableWidget.setItem(0, 1, newItem)

        newItem = QTableWidgetItem("Зачеты")
        newItem.setFont(self.font)
        newItem.setTextAlignment(QtCore.Qt.AlignCenter)
        self.ui.tableWidget.setItem(0, 2, newItem)

        newItem = QTableWidgetItem("Администратор доходов бюджета")
        newItem.setFont(self.font)
        newItem.setTextAlignment(QtCore.Qt.AlignCenter)
        self.ui.tableWidget.setItem(0, 3, newItem)

        newItem = QTableWidgetItem("Наименование бюджета")
        newItem.setFont(self.font)
        newItem.setTextAlignment(QtCore.Qt.AlignCenter)
        self.ui.tableWidget.setItem(0, 4, newItem)

    def ExportResultsToExcel(self):
        wb = Workbook()
        filename, _ = QFileDialog.getSaveFileName(self, 'Сохранить файл', '', ".xlsx(*.xlsx)")
        if filename in "('', '')":
            self.ui.statusbar.showMessage('Не выбран файл для сохранения.')
        else:
            try:
                self.sheet = wb.active
                row = 1
                col = 1
                wrap_alignment = Alignment(wrap_text=True)
                headerfont = Font(bold=True)
                for i in range(self.ui.tableWidget.rowCount()):
                    for x in range(self.ui.tableWidget.columnCount()):
                        try:
                            text = str(self.ui.tableWidget.item(i, x).text())
                            self.sheet.cell(row=row, column=col).value = text
                            self.sheet.cell(row=row, column=col).alignment = wrap_alignment
                            if row == 1:
                                self.sheet.cell(row=row, column=col).font = headerfont
                            col += 1
                        except AttributeError:
                            col += 1

                    row += 1
                    col = 1
                self.sheet.column_dimensions['A'].width = 20
                self.sheet.column_dimensions['B'].width = 20
                self.sheet.column_dimensions['C'].width = 20
                self.sheet.column_dimensions['D'].width = 75
                self.sheet.column_dimensions['E'].width = 75

                wb.save(filename=filename)
            except:
                self.ui.statusbar.showMessage('Ошибка при сохранении файла.')

    def OpenAbout(self):
        if (self.aboutForm != None):
            self.aboutForm.close()
        self.aboutForm = AboutWindows()
        self.aboutForm.show()

    @QtCore.pyqtSlot(list,bool)
    def appendText(self, list, _bool):
        if _bool:
            self.ui.statusbar.showMessage('Отчет проверен! Есть неисполненные поручения!')
            self.ui.tableWidget.setRowCount(len(list)+1)
            i = 1
            for ls in list:
                newItem = QTableWidgetItem(ls.licevoy)
                newItem.setFont(self.font_t)
                self.ui.tableWidget.setItem(i, 0, newItem)

                newItem = QTableWidgetItem(ls.vozvraty)
                newItem.setFont(self.font_t)
                self.ui.tableWidget.setItem(i, 1, newItem)

                newItem = QTableWidgetItem(ls.zachety)
                newItem.setFont(self.font_t)
                self.ui.tableWidget.setItem(i, 2, newItem)

                newItem = QTableWidgetItem(ls.dohody_admin)
                newItem.setFont(self.font_t)
                self.ui.tableWidget.setItem(i, 3, newItem)

                newItem = QTableWidgetItem(ls.budget)
                newItem.setFont(self.font_t)
                self.ui.tableWidget.setItem(i, 4, newItem)
                i = i+1
        else:
            self.ui.statusbar.showMessage('Отчет проверен! Неисполненных поручений нет!')